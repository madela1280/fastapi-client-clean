from fastapi import FastAPI, Request
from fastapi.responses import JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
import uvicorn
import re
from datetime import datetime

app = FastAPI()

# ✅ 반드시 FastAPI(app) 선언 바로 아래에 있어야 함
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],          # 프론트엔드 HTML에서 접근 허용
    allow_credentials=True,
    allow_methods=["*"],          # GET, POST, OPTIONS 등 모두 허용
    allow_headers=["*"],          # 모든 헤더 허용
)

@app.post("/get-user-info")
async def get_user_info(request: Request):
    body = await request.json()
    phone_input = body.get("queryResult", {}).get("parameters", {}).get("phone-number", "")

    # ✅ 리스트일 경우 첫 번째 값만 사용
    if isinstance(phone_input, list) and phone_input:
        phone_input = phone_input[0]

    # 숫자만 추출
    digits = re.sub(r'[^0-9]', '', phone_input)
    if len(digits) == 11:
        formatted_input = f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    else:
        formatted_input = phone_input

    print("📱 변환된 전화번호:", formatted_input)

    excel_path = "종합대여장부_자동화용.xlsx"
    sheet_name = "통합관리"

    try:
        wb = load_workbook(filename=excel_path, data_only=True)
        ws = wb[sheet_name]

        result = None

        for row in ws.iter_rows(min_row=2, values_only=True):
            tel1 = str(row[9])   # 연락처1
            tel2 = str(row[10])  # 연락처2
            returned = row[16]   # 반납완료일

            if formatted_input in [tel1, tel2] and not returned:
                name = row[8]     # 수취인명
                start = row[13]
                end = row[14]

                start_str = start.strftime("%Y.%m.%d") if isinstance(start, datetime) else str(start)
                end_str = end.strftime("%Y.%m.%d") if isinstance(end, datetime) else str(end)

                result = f"📦 대여자명: {name}\n📅 대여시작일: {start_str}\n⏳ 대여종료일: {end_str}"
                break

        if not result:
            result = "고객 정보를 찾을 수 없습니다.\n대여 시 등록한 정확한 전화번호를 입력해 주세요."

        return JSONResponse(content={"fulfillmentText": result})

    except Exception as e:
        print("❌ 오류:", str(e))
        return JSONResponse(content={"fulfillmentText": "시스템 오류가 발생했습니다. 잠시 후 다시 시도해 주세요."})

if __name__ == "__main__":
    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)
